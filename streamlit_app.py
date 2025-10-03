# streamlit_app.py
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import textwrap
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

# Import your logic
from mongo_connector import fetch_data_for_timeframe
from dgr_generator import generate_dgr_report

st.set_page_config(layout="wide", page_title="DGR Report Generator")
st.title("☀️ Tuljapur Solar Plant DGR Report Generator")

# --- Sidebar: Company Logo ---
enrich_logo_path = r"enrich_logo.png"
logo_path = r"Logo.png"
try:
    st.sidebar.image(enrich_logo_path, width=150)
except:
    st.sidebar.warning("Logo not found at provided path.")

# --- Sidebar: Report parameters ---
st.sidebar.header("Report Parameters")
report_type = st.sidebar.selectbox(
    "Select Report Type",
    ["Daily", "Weekly", "Monthly", "Custom Range"]
)

# Initialize dates
today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
start_date = today
end_date = today

# --- Date selection logic ---
if report_type == "Daily":
    selected_date = st.sidebar.date_input("Select Date", today - timedelta(days=1))
    start_date = datetime(selected_date.year, selected_date.month, selected_date.day)
    end_date = start_date
elif report_type == "Weekly":
    selected_date = st.sidebar.date_input("Select Week End Date", today - timedelta(days=1))
    end_date = datetime(selected_date.year, selected_date.month, selected_date.day)
    start_date = end_date - timedelta(days=6)
elif report_type == "Monthly":
    selected_month = st.sidebar.date_input("Select Month", today - timedelta(days=30))
    start_date = datetime(selected_month.year, selected_month.month, 1)
    next_month = start_date.replace(day=28) + timedelta(days=4)
    end_date = next_month - timedelta(days=next_month.day)
elif report_type == "Custom Range":
    start_d = st.sidebar.date_input("Start Date", today - timedelta(days=7))
    end_d = st.sidebar.date_input("End Date", today - timedelta(days=1))
    start_date = datetime(start_d.year, start_d.month, start_d.day)
    end_date = datetime(end_d.year, end_d.month, end_d.day)

st.sidebar.info(f"Report Period:\n{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

# --- Initialize session state ---
if "raw_db_data" not in st.session_state:
    st.session_state.raw_db_data = None
if "final_report_df" not in st.session_state:
    st.session_state.final_report_df = None
if "pdf_buffer" not in st.session_state:
    st.session_state.pdf_buffer = None
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None
if "csv_buffer" not in st.session_state:
    st.session_state.csv_buffer = None

# --- PDF generator ---
def create_professional_pdf(df, start_date, end_date, logo_path=None):
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib import pyplot as plt
    import textwrap
    import numpy as np
    from io import BytesIO

    buffer = BytesIO()
    rows_per_page = 25
    total_pages = (len(df) // rows_per_page) + 1

    df_wrapped = df.copy()
    if 'Remarks of the day' in df_wrapped.columns:
        df_wrapped['Remarks of the day'] = df_wrapped['Remarks of the day'].apply(
            lambda x: textwrap.fill(str(x), width=20)
        )

    with PdfPages(buffer) as pdf:
        for i, start in enumerate(range(0, len(df_wrapped), rows_per_page), 1):
            end = start + rows_per_page
            chunk = df_wrapped.iloc[start:end]
            fig_height = max(3, len(chunk)*0.35 + 3)
            fig, ax = plt.subplots(figsize=(16, fig_height))
            ax.axis('off')

            # --- DGR Statement top-left ---
            ax.text(0.0, 1.05,
                    f"DGR Generated for: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}",
                    transform=ax.transAxes, fontsize=11, ha='left', va='top', fontweight='bold')

            # --- Logo on left ---
            if logo_path:
                try:
                    img = plt.imread(logo_path)
                    ax.imshow(img, 
                              extent=(0.01, 0.08, 0.95, 1.03), 
                              zorder=10, 
                              aspect='auto')  # small logo top-left
                except:
                    pass

            # --- Company info next to logo ---
            company_info = [
                "Imagicaaworld Entertainment Ltd",
                "AC 5.5 MW / DC 8.0 MW Solar PV Project",
                "Tuljapur Site",
                "Comm. Date: 30 Jul 2024"
            ]
            for idx, line in enumerate(company_info):
                ax.text(0.09, 0.95 - idx*0.03, line, transform=ax.transAxes, fontsize=10, ha='left', va='top')

            # --- Footer: page number ---
            ax.text(1, -0.05, f"Page {i} of {total_pages}", transform=ax.transAxes, fontsize=9, ha='right')

            # --- Table ---
            table = ax.table(
                cellText=chunk.values,
                colLabels=chunk.columns,
                cellLoc='center',
                loc='center'
            )
            table.auto_set_font_size(False)
            table.set_fontsize(7)
            for j, col in enumerate(chunk.columns):
                max_len = max(chunk[col].astype(str).apply(len).max(), len(col))
                for k in range(len(chunk)):
                    table._cells[(k+1,j)].set_width(max_len*0.012)
                table._cells[(0,j)].set_width(max_len*0.012)

            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)

    buffer.seek(0)
    return buffer

# --- Excel generator ---
def generate_excel(df, start_date, end_date, logo_path=None):
    """
    Generates Excel with:
    - DGR statement merged across 10 columns
    - Logo on top-left
    - Company info next to logo without leaving gaps
    - DGR data below
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=7, sheet_name="DGR Report")
        workbook = writer.book
        worksheet = writer.sheets["DGR Report"]

        # --- DGR Statement: Merge A1:J1 ---
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        worksheet.cell(row=1, column=1, value=f"DGR Generated for: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}")
        worksheet.cell(row=1, column=1).font = Font(bold=True)
        worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

        # --- Logo ---
        if logo_path:
            try:
                logo = Image(logo_path)
                logo.height = 60
                logo.width = 60
                worksheet.add_image(logo, "A2")  # Top-left
            except Exception as e:
                print(f"Logo not added: {e}")

        # --- Company info next to logo without empty column ---
        header_info = [
            ["Name:", "Imagicaaworld Entertainment Ltd"],
            ["Capacity:", "AC 5.5 MW / DC 8.0 MW Solar PV Project"],
            ["Site Name:", "Tuljapur Site"],
            ["Comm. Date:", "30 Jul 2024"]
        ]
        start_row = 2
        start_col = 2  # Start immediately next to logo
        for r, row in enumerate(header_info, start=start_row):
            for c, value in enumerate(row, start=start_col):
                cell = worksheet.cell(row=r, column=c, value=value)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")

    output.seek(0)
    return output

# --- Main app logic ---
if st.button("Generate Report"):
    with st.spinner(f"Fetching and processing data from {start_date.date()} to {end_date.date()}..."):
        # Fetch data
        st.session_state.raw_db_data = fetch_data_for_timeframe(start_date, end_date)
        
        if not st.session_state.raw_db_data:
            st.warning("⚠️ No data found for the selected period.")
        else:
            # Generate DGR report DataFrame
            st.session_state.final_report_df = generate_dgr_report(st.session_state.raw_db_data)
            
            # CSV buffer
            st.session_state.csv_buffer = st.session_state.final_report_df.to_csv(index=False).encode('utf-8')
            
            # PDF buffer (pass start_date and end_date)
            st.session_state.pdf_buffer = create_professional_pdf(
                st.session_state.final_report_df,
                start_date,
                end_date,
                logo_path=logo_path
            )
            
            # Excel buffer (pass start_date and end_date if needed inside generator)
            st.session_state.excel_buffer = generate_excel(
                st.session_state.final_report_df,
                start_date=start_date,
                end_date=end_date,
                logo_path=logo_path
            )
            
            st.success("✅ Report Generated Successfully!")

# --- Display & Download Buttons ---
if st.session_state.final_report_df is not None:
    st.dataframe(st.session_state.final_report_df)

    st.download_button(
        label="📥 Download CSV",
        data=st.session_state.csv_buffer,
        file_name=f"DGR_Report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv",
        mime="text/csv",
        key="csv_download"
    )

    st.download_button(
        label="📥 Download PDF",
        data=st.session_state.pdf_buffer,
        file_name=f"DGR_Report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf",
        mime="application/pdf",
        key="pdf_download"
    )

    st.download_button(
        label="📥 Download Excel",
        data=st.session_state.excel_buffer,
        file_name=f"DGR_Report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="excel_download"
    )

    # Optional: show raw data
    with st.expander("Show Raw Fetched Data"):
        st.write(pd.DataFrame(st.session_state.raw_db_data))
